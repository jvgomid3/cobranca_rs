r"""
Atualiza a planilha Controle_Cobranca_R&S.xlsx a partir do arquivo RP_COBRANCAS.TXT.

Regras implementadas (conforme você descreveu):
- Lê RP_COBRANCAS.TXT (na pasta S:\...\HRS1)
- Para cada RP (coluna 1), se já existir na planilha (coluna "ID Vaga"), pula.
- Se não existir, adiciona uma nova linha preenchendo:
  ID Vaga      = col 1 do RP
  Nome do Aprovado = col 8 do RP
  Centro cst   = col 6 do RP
  Cargo        = (placeholder: tenta buscar via CARGOS_RBLA.TXT; se não achar, fica em branco)
  Índice       = (placeholder: derivado do Cargo via dicionário; se não achar, fica em branco)
  Qtd          = 1
  Status       = col 5 (E) do RP: 2 -> "2 - Fechada", 3 -> "3 - Cancelada" (senão mantém o valor bruto)
  Mês/Ano      = col 11 do RP: "01.03.2026" -> "Março/2026"
  Faturar?     = col 12 do RP
  Número Cobrança = em branco (não foi especificado)
"""

from __future__ import annotations

import os
import sys
import shutil
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =========================
# CONFIG (ajuste se quiser)
# =========================
RP_FILE_PATH = r"S:\HRC\hrs-br\hrs-br\Inter_Setor\BOTs\Cobrancas\HRS1\RP_COBRANCAS.TXT"
CARGOS_FILE_PATH = r"S:\HRC\hrs-br\hrs-br\Inter_Setor\BOTs\Cobrancas\HRS1\CARGOS_RBLA.TXT"
XLSX_PATH = r"S:\HRC\hrs-br\hrs-br\Inter_Setor\BOTs\Cobranca_RS\Controle_Cobranca_R&S.xlsx"

SHEET_NAME = None  # None = primeira aba
MAKE_BACKUP = False

# Mapeamento do "Índice" por Cargo (PLACEHOLDER)
# Depois você substitui isso pela regra real.
INDICE_POR_CARGO: Dict[str, str] = {
    # "Analista": "X",
    # "Coordenador": "Y",
}

# Se CARGOS_RBLA.TXT tiver mapeamento de ID Vaga -> Cargo, vamos tentar pegar.
# Ajuste abaixo caso você descubra outra posição/coluna no arquivo.
CARGOS_HINTS = {
    "id_vaga_col": 1,   # 1-indexed
    "cargo_col": 2,     # 1-indexed (chute comum: ID;Cargo;...)
}


# ==========
# LOGGING
# ==========
def ensure_standard_streams() -> None:
    if sys.stdin is None:
        sys.stdin = open(os.devnull, "r", encoding="utf-8")
    if sys.stdout is None:
        sys.stdout = open(os.devnull, "w", encoding="utf-8")
    if sys.stderr is None:
        sys.stderr = open(os.devnull, "w", encoding="utf-8")


def get_log_file_path() -> str:
    log_dir = os.environ.get("TEMP") or os.getcwd()
    return os.path.join(log_dir, "cobranca_rs_atualizacao.log")


def setup_logger() -> logging.Logger:
    ensure_standard_streams()

    logger = logging.getLogger("cobranca_rs")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    logger.propagate = False

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    fh = logging.FileHandler(get_log_file_path(), encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


logger = setup_logger()


# ==========
# HELPERS
# ==========
PT_MONTHS = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}


def safe_strip(s: Optional[str]) -> str:
    return (s or "").strip()


def detect_delimiter(sample_line: str) -> str:
    # Tentativa simples para TXT corporativo:
    candidates = [";", "|", "\t", ","]
    counts = {d: sample_line.count(d) for d in candidates}
    best = max(counts, key=counts.get)
    # se nenhum delimitador aparece, cai no ';' por padrão
    return best if counts[best] > 0 else ";"


def read_text_lines(path: str) -> List[str]:
    # tenta alguns encodings comuns em ambiente Windows corporativo
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    last_err = None
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                return [line.rstrip("\n") for line in f]
        except Exception as e:
            last_err = e
    # fallback permissivo
    with open(path, "r", encoding="latin-1", errors="replace") as f:
        return [line.rstrip("\n") for line in f]


def parse_dd_mm_yyyy_dot(s: str) -> Optional[datetime]:
    s = safe_strip(s)
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def format_mes_ano_pt(dt: datetime) -> str:
    return f"{PT_MONTHS[dt.month]}/{dt.year}"


def status_label(raw: str) -> str:
    raw = safe_strip(raw)
    # Remove zeros à esquerda para comparação
    if raw in ("2", "02"):
        return "2 - Fechada"
    if raw in ("3", "03"):
        return "3 - Cancelada"
    return raw  # não especificado: mantém como veio


def format_id_vaga(raw: str) -> str:
    """Remove zeros à esquerda do ID da vaga."""
    raw = safe_strip(raw)
    if not raw:
        return raw
    # Remove zeros à esquerda, mas mantém se for apenas "0"
    return raw.lstrip('0') or '0'


@dataclass
class RPRow:
    id_vaga: str
    nome_aprovado: str
    centro_cst: str
    tipo_vaga: str  # coluna 4: G2, V2, CH, MS, MN, etc.
    cargo_id: str
    is_pcd: bool  # coluna 9: 1=PCD, 2 ou vazio=não PCD
    status: str
    mes_ano: str
    faturar: str


def parse_rp_rows(rp_path: str) -> Tuple[List[RPRow], str]:
    lines = [ln for ln in read_text_lines(rp_path) if safe_strip(ln)]
    if not lines:
        return [], ";"

    delim = detect_delimiter(lines[0])
    rows: List[RPRow] = []

    for i, ln in enumerate(lines, start=1):
        parts = [p.strip() for p in ln.split(delim)]
        # Precisamos no mínimo até coluna 12
        if len(parts) < 12:
            logger.warning(f"Linha {i} ignorada (colunas insuficientes: {len(parts)}): {ln[:120]}")
            continue

        # colunas 1-indexed
        id_vaga = format_id_vaga(parts[0])
        tipo_vaga = safe_strip(parts[3])  # col 4: tipo de vaga (G2, V2, CH, etc.)
        centro_cst = safe_strip(parts[5])
        cargo_id = safe_strip(parts[6])  # col 7: para buscar Cargo SAP
        nome_aprovado = safe_strip(parts[7])
        pcd_flag = safe_strip(parts[8])  # col 9: 1=PCD, 2 ou vazio=não PCD
        is_pcd = (pcd_flag == "1")
        status_raw = safe_strip(parts[4])
        dt_raw = safe_strip(parts[10])
        faturar = safe_strip(parts[11])

        dt = parse_dd_mm_yyyy_dot(dt_raw)
        mes_ano = format_mes_ano_pt(dt) if dt else safe_strip(dt_raw)

        if not id_vaga:
            logger.warning(f"Linha {i} ignorada (ID Vaga vazio).")
            continue

        rows.append(
            RPRow(
                id_vaga=id_vaga,
                nome_aprovado=nome_aprovado,
                centro_cst=centro_cst,
                tipo_vaga=tipo_vaga,
                cargo_id=cargo_id,
                is_pcd=is_pcd,
                status=status_label(status_raw),
                mes_ano=mes_ano,
                faturar=faturar,
            )
        )

    return rows, delim


def determine_cargo_catalogo_indice(tipo_vaga: str, cargo_sap: str, is_pcd: bool = False) -> Tuple[str, str]:
    """
    Determina Cargo Catálogo e Índice baseado no tipo de vaga (coluna 4 do RP),
    Cargo SAP e se é vaga PCD (coluna 9 do RP).
    Retorna (cargo_catalogo, indice).
    """
    tipo_vaga = tipo_vaga.upper().strip()
    cargo_sap_upper = cargo_sap.upper().strip()

    # ===== REGRAS PCD (Affirmative positions) - Prioridade maior =====
    if is_pcd:
        # Regra PCD especial: Cargo começando com "Tecnico" (sempre Jr/HI, mesmo com Pl ou Sr)
        if cargo_sap_upper.startswith("TECNICO"):
            return "Affirmative position - MN Jr. / HI", "HRSR18"
        
        # Regra PCD 1: MS ou (MN E Cargo SAP contém "Sr")
        if tipo_vaga == "MS" or (tipo_vaga == "MN" and "SR" in cargo_sap_upper):
            return "Affirmative position - MN Sr.", "HRSR16"
        
        # Regra PCD 2: MN E Cargo SAP contém "Pl"
        if tipo_vaga == "MN" and "PL" in cargo_sap_upper:
            return "Affirmative position - MN Pl.", "HRSR17"
        
        # Regra PCD 3: HN ou MN com cargo contendo Assist* ou Jr
        if tipo_vaga == "HN" or (
            tipo_vaga == "MN"
            and (
                "ASSIST" in cargo_sap_upper
                or "JR" in cargo_sap_upper
            )
        ):
            return "Affirmative position - MN Jr. / HI", "HRSR18"
        
        # Regra PCD 4: EB ou EU
        if tipo_vaga in ("EB", "EU"):
            return "Affirmative position - Intern", "HRSR19"
        
        # Regra PCD 5: HD ou HA
        if tipo_vaga in ("HD", "HA"):
            return "Affirmative position - HD / Apprentice (HA)", "HRSR20"
    
    # ===== REGRAS BASEADAS EM CARGO SAP (Prioridade sobre regras de tipo_vaga) =====
    # Regra Cargo SAP especial: Técnico (sempre Jr/HI, mesmo com Pl ou Sr)
    if cargo_sap_upper.startswith("TECNICO"):
        return "MN Ass, Jr / HI", "HRSR06"
    
    # Regra Cargo SAP 1: Supervisor
    if "SUPERVISOR" in cargo_sap_upper:
        return "R&S - MN Sr.", "HRSR04"
    
    # Regra Cargo SAP 2: Líder
    if "LIDER" in cargo_sap_upper or "Leader" in cargo_sap_upper.title():
        return "R&S - MN Jr. / HI", "HRSR06"
    
    # ===== REGRAS NORMAIS (Não PCD) =====
    # Regra 1: G2 ou V2
    if tipo_vaga in ("G2", "V2"):
        return "SL2", "HRSR01"

    # Regra 2: G1 ou V1
    if tipo_vaga in ("G1", "V1"):
        return "SL1", "HRSR02"

    # Regra 3: CH
    if tipo_vaga == "CH":
        return "SLR", "HRSR03"

    # Regra 4: MS ou (MN E Cargo SAP contém "Sr")
    if tipo_vaga == "MS" or (tipo_vaga == "MN" and "SR" in cargo_sap_upper):
        return "MN Sr.", "HRSR04"

    # Regra 5: MN E Cargo SAP contém "Pl"
    if tipo_vaga == "MN" and "PL" in cargo_sap_upper:
        return "MN Pl.", "HRSR05"

    # Regra 6: HN ou MN com cargo contendo Assist* ou Jr
    if tipo_vaga == "HN" or (
        tipo_vaga == "MN"
        and (
            "ASSIST" in cargo_sap_upper
            or "JR" in cargo_sap_upper
        )
    ):
        return "MN Ass, Jr / HI", "HRSR06"

    # Regra 7: HD ou HA
    if tipo_vaga in ("HD", "HA"):
        return "HD / Apprentice (HA)", "HRSR08"

    # Regra 8: EB ou EU -> Intern
    # Nota: Havia duplicação na regra (também HRSR16), usando apenas Intern
    if tipo_vaga in ("EB", "EU"):
        return "Intern", "HRSR09"

    # Se não se encaixar em nenhuma regra, retorna vazio
    return "", ""


def load_cargos_sap_mapping(cargos_path: str) -> Dict[str, str]:
    """
    Monta um dict {CARGO_ID: CARGO_SAP} a partir de CARGOS_RBLA.TXT.
    - Chave: coluna 2 do CARGOS_RBLA.TXT
    - Valor: coluna 4 do CARGOS_RBLA.TXT
    """
    if not os.path.exists(cargos_path):
        logger.warning(f"CARGOS_RBLA.TXT não encontrado em: {cargos_path}")
        return {}

    lines = [ln for ln in read_text_lines(cargos_path) if safe_strip(ln)]
    if not lines:
        return {}

    delim = detect_delimiter(lines[0])
    mapping: Dict[str, str] = {}

    for i, ln in enumerate(lines, start=1):
        parts = [p.strip() for p in ln.split(delim)]
        # Precisamos pelo menos até a coluna 4 (index 3)
        if len(parts) < 4:
            continue
        # Coluna 2 (index 1) -> chave
        # Coluna 4 (index 3) -> valor (Cargo SAP)
        key = safe_strip(parts[1])
        val = safe_strip(parts[3])
        if key and val:
            mapping[key] = val

    logger.info(f"Mapeamento Cargo SAP carregado: {len(mapping)} registros (delim='{delim}')")
    return mapping


def ensure_headers(ws: Worksheet, headers: List[str]) -> Dict[str, int]:
    """
    Garante que a linha 1 tem os headers esperados.
    Retorna dict header->col_index (1-indexed).
    """
    existing = [safe_strip(ws.cell(1, c).value if ws.cell(1, c).value is not None else "") for c in range(1, len(headers) + 1)]
    if any(h != headers[idx] for idx, h in enumerate(existing)):
        # Se a planilha estiver vazia na primeira linha, escreve.
        if all(h == "" for h in existing):
            for c, h in enumerate(headers, start=1):
                ws.cell(1, c).value = h
            logger.info("Cabeçalhos criados na planilha.")
        else:
            logger.warning(
                "Cabeçalhos diferentes do esperado. Vou tentar operar mesmo assim "
                "mapeando pelas posições padrão."
            )

    header_to_col = {headers[i]: i + 1 for i in range(len(headers))}
    return header_to_col


def get_existing_ids(ws: Worksheet, id_col: int) -> set[str]:
    existing = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, id_col).value
        v = safe_strip(str(v)) if v is not None else ""
        if v:
            existing.add(v)
    return existing


def append_row(ws: Worksheet, values_by_col: Dict[int, str]) -> int:
    new_r = ws.max_row + 1
    for col, val in values_by_col.items():
        ws.cell(new_r, col).value = val
    return new_r


def make_backup(xlsx_path: str) -> Optional[str]:
    if not os.path.exists(xlsx_path):
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{xlsx_path}.backup_{ts}"
    shutil.copy2(xlsx_path, backup_path)
    return backup_path


def missing_path_message(label: str, path: str) -> str:
    msg = f"{label} não encontrado: {path}"
    if len(path) >= 2 and path[1] == ":":
        msg += (
            " | No Task Scheduler, confirme se a unidade mapeada está disponível para a conta da tarefa "
            "ou troque para caminho UNC."
        )
    return msg


def main() -> int:
    if not os.path.exists(RP_FILE_PATH):
        logger.error(missing_path_message("RP_COBRANCAS.TXT", RP_FILE_PATH))
        return 1

    if not os.path.exists(XLSX_PATH):
        logger.error(missing_path_message("Planilha", XLSX_PATH))
        return 1

    if MAKE_BACKUP:
        bkp = make_backup(XLSX_PATH)
        if bkp:
            logger.info(f"Backup criado: {bkp}")

    rp_rows, rp_delim = parse_rp_rows(RP_FILE_PATH)
    logger.info(f"RPs lidas: {len(rp_rows)} (delim='{rp_delim}')")

    cargos_sap_map = load_cargos_sap_mapping(CARGOS_FILE_PATH)

    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.worksheets[0]

    headers = [
        "ID Vaga",
        "Nome do Aprovado",
        "Centro Custo",
        "Subgrupo",
        "Cargo SAP",
        "Cargo Catálogo",
        "Índice",
        "Qtd",
        "Status",
        "Mês/Ano",
        "Faturar?",
        "Número Cobrança",
    ]
    hcol = ensure_headers(ws, headers)

    existing_ids = get_existing_ids(ws, hcol["ID Vaga"])
    logger.info(f"IDs já existentes na planilha: {len(existing_ids)}")

    inserted = 0
    skipped = 0

    for rp in rp_rows:
        if rp.id_vaga in existing_ids:
            skipped += 1
            continue

        cargo_sap = cargos_sap_map.get(rp.cargo_id, "")
        cargo_catalogo, indice = determine_cargo_catalogo_indice(rp.tipo_vaga, cargo_sap, rp.is_pcd)
        
        # Se Faturar? = Não, então Número Cobrança = Não Cobrar
        numero_cobranca = "Não Cobrar" if rp.faturar.strip().lower() == "não" else ""

        values = {
            hcol["ID Vaga"]: rp.id_vaga,
            hcol["Nome do Aprovado"]: rp.nome_aprovado,
            hcol["Centro Custo"]: rp.centro_cst,
            hcol["Subgrupo"]: rp.tipo_vaga,
            hcol["Cargo SAP"]: cargo_sap,
            hcol["Cargo Catálogo"]: cargo_catalogo,  # placeholder
            hcol["Índice"]: indice,
            hcol["Qtd"]: "1",
            hcol["Status"]: rp.status,
            hcol["Mês/Ano"]: rp.mes_ano,
            hcol["Faturar?"]: rp.faturar,
            hcol["Número Cobrança"]: numero_cobranca,
        }

        new_row = append_row(ws, values)
        existing_ids.add(rp.id_vaga)
        inserted += 1

        if inserted % 200 == 0:
            logger.info(f"Inseridas {inserted} linhas... (última linha: {new_row})")

    wb.save(XLSX_PATH)
    logger.info(f"Concluído. Inseridas: {inserted} | Puladas (já existiam): {skipped}")
    logger.info(f"Arquivo atualizado: {XLSX_PATH}")

    if inserted == 0:
        logger.warning("Nenhuma linha nova foi inserida.")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:
        logger.exception("Falha não tratada durante a atualização.")
        raise