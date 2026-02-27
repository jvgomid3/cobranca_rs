# Importa a biblioteca customtkinter, versão personalizada do Tkinter
import customtkinter as ctk
# Importa messagebox do Tkinter para exibir mensagens de alerta e informação
from tkinter import messagebox
# Importa load_workbook da openpyxl para manipular arquivos Excel
from openpyxl import load_workbook
# Importa Playwright para automação de navegador
from playwright.sync_api import sync_playwright
# Importa pyperclip para copiar conteúdo para o clipboard
import pyperclip
# Importa biblioteca os para manipulação de arquivos e caminhos
import os
# Importa expressões regulares para buscar padrões em texto
import re
# Importa win32com.client para controlar o Outlook via COM
import win32com.client as win32
# Importa datetime para manipulação de datas
from datetime import datetime

# -------------------- Configurações SAP Web -------------------- #
SAP_WEB_URL = "https://qs0.wdisp.bosch.com/sap/bc/gui/sap/its/webgui#"
STORAGE_STATE_PATH = "sap_session.json"

# -------------------- Função para converter mês/ano -------------------- #
def mes_ano_para_formato_curto(mes_ano):
    try:
        mes_nome, ano = mes_ano.split("/")
        mes_nome = mes_nome.strip().lower()
        meses = {
            "janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4,
            "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
            "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12
        }
        mes_num = meses.get(mes_nome, 0)
        ano_curto = str(ano.strip())[-2:]
        return f"{mes_num:02d}.{ano_curto}" if mes_num else mes_ano
    except Exception as e:
        print(f"Erro ao converter mês/ano: {e}")
        return mes_ano

# -------------------- Atualizar planilha -------------------- #
def atualizar_planilha(encontrados, numero):
    try:
        wb = load_workbook(caminho_excel)
        ws = wb.active
        cabecalho = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_id = cabecalho.index("ID Vaga")
        idx_mes = cabecalho.index("Mês/Ano")
        idx_cobranca = cabecalho.index("Número Cobrança")

        for row in ws.iter_rows(min_row=2):
            id_vaga = row[idx_id].value
            mes_ano = row[idx_mes].value
            for d in encontrados:
                if d["id"] == id_vaga and d["mes"] == mes_ano:
                    row[idx_cobranca].value = numero

        wb.save(caminho_excel)
        wb.close()
        print(f"Planilha atualizada com o número {numero}")

    except Exception as e:
        messagebox.showerror("Erro Excel", f"Falha ao atualizar planilha:\n{e}")

# -------------------- Cancelar Cobrança -------------------- #
def cancelar_cobranca_selecionados():
    """Marca 'Não Cobrar' nos itens selecionados tanto no Excel quanto na interface"""
    # Filtrar apenas os itens selecionados via checkbox
    selecionados = [item["data"] for item in checkbox_vars if item["var"].get()]
    
    if not selecionados:
        messagebox.showwarning("Atenção", "Nenhum item selecionado.")
        return
    
    resposta = messagebox.askyesno(
        "Confirmação",
        f"Deseja marcar {len(selecionados)} registro(s) como 'Não Cobrar'?"
    )
    
    if not resposta:
        return
    
    try:
        # Atualizar planilha Excel
        wb = load_workbook(caminho_excel)
        ws = wb.active
        cabecalho = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        idx_id = cabecalho.index("ID Vaga")
        idx_mes = cabecalho.index("Mês/Ano")
        idx_cobranca = cabecalho.index("Número Cobrança")
        
        for row in ws.iter_rows(min_row=2):
            id_vaga = row[idx_id].value
            mes_ano = row[idx_mes].value
            for d in selecionados:
                if d["id"] == id_vaga and d["mes"] == mes_ano:
                    row[idx_cobranca].value = "Não Cobrar"
                    # Atualizar também nos dados em memória
                    d["cobranca"] = "Não Cobrar"
        
        wb.save(caminho_excel)
        wb.close()
        print(f"{len(selecionados)} registro(s) marcado(s) como 'Não Cobrar'")
        
        # Atualizar a tabela na interface
        atualizar_tabela()
        
        messagebox.showinfo("Sucesso", f"{len(selecionados)} registro(s) marcado(s) como 'Não Cobrar'")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao cancelar cobrança:\n{e}")

# -------------------- Abrir SAP Web -------------------- #
def abrir_sap_web(mes_ano, encontrados):
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=False,
                channel="chrome",
                args=[
                    "--enable-features=ClipboardReadWrite",
                    "--disable-features=IsolateOrigins,site-per-process",
                    "--unsafely-treat-insecure-origin-as-secure=https://qs0.wdisp.bosch.com"
                ]
            )

            if os.path.exists(STORAGE_STATE_PATH):
                context = browser.new_context(
                    storage_state=STORAGE_STATE_PATH,
                    permissions=["clipboard-read", "clipboard-write"]
                )
            else:
                context = browser.new_context(
                    permissions=["clipboard-read", "clipboard-write"]
                )

            page = context.new_page()
            page.set_viewport_size({"width": 1366, "height": 768})
            page.goto(SAP_WEB_URL)
            page.evaluate("document.body.style.zoom='80%'")
            
            # Verificação de login
            try:
                login_user = page.locator('//*[@id="sap-user"]')
                if login_user.is_visible():
                    print("Tela de login detectada — aguarde login manual...")
                    page.wait_for_selector('//*[@id="sap-password"]', timeout=15000)
                    page.wait_for_function("document.querySelector('#sap-user') === null", timeout=120000)
                    print("Login realizado, continuando...")
            except Exception:
                pass
            
            page.wait_for_timeout(5000)

            # Verificação do botão de continuar
            try:
                cont_btn = page.locator('//*[@id="SYSTEM_MESSAGE_CONTINUE_BUTTON"]')
                if cont_btn.is_visible():
                    print("Botão de avançar detectado — será clicado...")
                    cont_btn.click()
                    page.wait_for_timeout(1500)
                else:
                    print("Botão de avançar não detectado — pulando para KB31N...")
            except Exception:
                print("Botão de avançar não detectado — pulando para KB31N...")

            # Aguarda a tela carregar completamente antes de inserir KB31N
            page.wait_for_timeout(1000)

            from playwright.sync_api import TimeoutError

            for tentativa in range(3):
                try:
                    page.wait_for_selector('//*[@id="ToolbarOkCode"]', timeout=120000)
                    page.wait_for_timeout(1000)  # Espera adicional antes de inserir
                    page.fill('//*[@id="ToolbarOkCode"]', '/nKB31N')
                    page.press('//*[@id="ToolbarOkCode"]', 'Enter')
                    page.wait_for_timeout(2000)
                    break
                except TimeoutError:
                    print(f"Tentativa {tentativa+1} falhou, tentando novamente...")

            page.wait_for_timeout(3000)

            # Verificação do campo empresa
            try:
                campo_empresa = page.locator('//*[@id="M1:46:1::0:21"]')
                if campo_empresa.is_visible():
                    campo_empresa.fill('0010')
                    campo_empresa.press('Enter')
                    page.wait_for_timeout(3000)
            except Exception as e:
                print(f"Campo empresa não encontrado. Pulando... ({e})")

            page.click('//*[@id="M0:46:2:1::13:35"]')
            page.wait_for_timeout(2000)
            
            # Clica no botão e preenche o campo de texto
            try:
                page.click('//*[@id="M0:46:1:1:2B256::4:12"]')
                page.wait_for_timeout(1000)
                
                # Pega o mês e ano atual
                agora = datetime.now()
                meses_pt = {
                    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho",
                    7: "Julho", 8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
                }
                mes_atual = meses_pt[agora.month]
                ano_atual = agora.year
                
                # Monta o texto completo
                texto_campo = f"Recuperação custo {mes_atual}/{ano_atual} HRS2-LA R&S"
                
                # Preenche o campo usando keyboard.type
                page.keyboard.type(texto_campo)
                print(f"Campo preenchido com: {texto_campo}")
                
            except Exception as e:
                print(f"Erro ao preencher campo: {e}")
            
            # Aguarda um momento
            page.wait_for_timeout(500)
            
            # ========== SALVAR E CAPTURAR NÚMERO (COMENTADO PARA TESTES) ========== #
            # numero_cobranca = None
            # try:
            #     print("Clicando em salvar...")
            #     page.click('//*[@id="M0:36::btn[11]"]')
            #     page.wait_for_timeout(2000)
            #     
            #     # Captura a mensagem da barra de status
            #     mensagem = page.inner_text('xpath=//*[@id="wnd[0]/sbar_msg-txt"]')
            #     print(f"Mensagem capturada: {mensagem}")
            #     
            #     # Extrai o número da mensagem
            #     match = re.search(r"\d+", mensagem)
            #     if match:
            #         numero_cobranca = match.group()
            #         print(f"Número de cobrança gerado: {numero_cobranca}")
            #     else:
            #         print("Nenhum número encontrado na mensagem!")
            #         
            # except Exception as e:
            #     print(f"Erro ao salvar e capturar número: {e}")
            # 
            # # Só continua se conseguiu capturar o número
            # if not numero_cobranca:
            #     print("ATENÇÃO: Não foi possível capturar o número de cobrança.")
            #     return
            # ======================================================================= #
            
            page.wait_for_timeout(1000)
            
            messagebox.showinfo("SAP Web", "Processo executado (sem salvar).\nVerifique a tela do SAP.")
            
            context.storage_state(path=STORAGE_STATE_PATH)
            # browser.close()  # Comentado para você verificar a tela
            # app.destroy()  # Comentado para você verificar a tela

    except Exception as e:
        messagebox.showerror("Erro SAP Web", f"Não foi possível seguir com o SAP Web.\n{e}")
        # app.destroy()  # Comentado para você verificar o erro

# -------------------- Excel - Carregar dados -------------------- #
caminho_excel = r"C:\Users\ajl8ca\Desktop\HRS_Projects_Dev\cobranca_r&s\Controle_Cobranca_R&S.xlsx"

def carregar_dados():
    try:
        wb = load_workbook(caminho_excel, read_only=True, data_only=True)
        ws = wb.active
        cabecalho = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        colunas_necessarias = ["ID Vaga", "Nome do Aprovado", "Centro cst", "Mês/Ano", "Índice", "Faturar?", "Status", "Número Cobrança"]
        for col in colunas_necessarias:
            if col not in cabecalho:
                raise ValueError(f"Coluna '{col}' não encontrada.")

        idx_id = cabecalho.index("ID Vaga")
        idx_nome = cabecalho.index("Nome do Aprovado")
        idx_centro = cabecalho.index("Centro cst")
        idx_mes = cabecalho.index("Mês/Ano")
        idx_indice = cabecalho.index("Índice")
        idx_faturar = cabecalho.index("Faturar?")
        idx_status = cabecalho.index("Status")
        idx_cobranca = cabecalho.index("Número Cobrança")

        dados = []
        meses_unicos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx_mes] is None:
                continue
            mes = str(row[idx_mes])
            if mes not in meses_unicos:
                meses_unicos.append(mes)
            dados.append({
                "mes": mes,
                "id": row[idx_id],
                "nome": row[idx_nome],
                "centro": row[idx_centro],
                "indice": row[idx_indice],
                "faturar": row[idx_faturar],
                "status": row[idx_status],
                "cobranca": row[idx_cobranca]
            })

        wb.close()
        return meses_unicos, dados

    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível ler o arquivo:\n{e}")
        return [], []

# -------------------- UI -------------------- #
meses_anos, dados_planilha = carregar_dados()

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

app = ctk.CTk()
app.title("Faturamento R&S")

ctk.CTkLabel(app, text="Faturamento R&S", font=("Arial", 18, "bold")).pack(pady=10)
ctk.CTkLabel(app, text="Selecione o Mês/Ano:").pack(pady=0)
combo = ctk.CTkComboBox(app, values=meses_anos, width=200, justify="center")
combo.pack(pady=5)

frame_confirmar = ctk.CTkFrame(app, fg_color="transparent", height=80)
frame_confirmar.pack(fill="x", pady=10)

btn_confirmar = ctk.CTkButton(frame_confirmar, text="Confirmar", width=150)
btn_confirmar.place(relx=0.5, rely=0.3, anchor="center")

btn_cancelar_cobranca = ctk.CTkButton(frame_confirmar, text="Cancelar Cobrança", width=150)
btn_cancelar_cobranca.place(relx=0.88, rely=0.3, anchor="center")

filtro_var = ctk.BooleanVar(value=False)
check_filtro = ctk.CTkCheckBox(frame_confirmar, text="Filtrar Vazias", variable=filtro_var)
check_filtro.place(relx=0.78, rely=0.7, anchor="center")

filtro_canceladas_var = ctk.BooleanVar(value=False)
check_filtro_canceladas = ctk.CTkCheckBox(frame_confirmar, text="Filtrar Canceladas", variable=filtro_canceladas_var)
check_filtro_canceladas.place(relx=0.88, rely=0.7, anchor="center")

scrollable_frame = ctk.CTkScrollableFrame(app)
scrollable_frame.pack(pady=10, padx=20, fill="both", expand=True)

# Lista para armazenar as variáveis de checkbox e dados associados
checkbox_vars = []

# Variável para checkbox "Selecionar Todas"
selecionar_todas_var = ctk.BooleanVar(value=True)

def toggle_selecionar_todas():
    """Marca ou desmarca todas as checkboxes das linhas"""
    estado = selecionar_todas_var.get()
    for item in checkbox_vars:
        item["var"].set(estado)

headers = ["", "ID Vaga", "Nome do Aprovado", "Centro de Custo", "Índice", "Faturar?", "Status", "Número Cobrança"]
for col, h in enumerate(headers):
    if col == 0:
        # Coluna de checkbox - com checkbox "Selecionar Todas"
        check_selecionar_todas = ctk.CTkCheckBox(
            scrollable_frame, 
            text="", 
            variable=selecionar_todas_var, 
            command=toggle_selecionar_todas,
            width=20
        )
        check_selecionar_todas.grid(row=0, column=col, padx=2, pady=5, sticky="nsew")
        scrollable_frame.grid_columnconfigure(col, minsize=40, weight=0)
    else:
        ctk.CTkLabel(scrollable_frame, text=h, font=("Arial", 12, "bold")).grid(
            row=0, column=col, padx=5, pady=5, sticky="nsew"
        )
        scrollable_frame.grid_columnconfigure(col, weight=1)

executar_btn = ctk.CTkButton(app, text="Executar", command=lambda: executar())
executar_btn.pack(pady=10)

# -------------------- Funções Auxiliares -------------------- #
def preparar_clipboard(dados_filtrados, mes_ano):
    linhas = []
    for d in dados_filtrados:
        centro = str(d["centro"])
        indice = d.get("indice", "HRSR26")
        faturar = d.get("faturar", "1")
        texto = str(d["id"])  # Usa o ID Vaga como texto
        linha = f"{centro}\t{indice}\t{faturar}\t{texto}"
        linhas.append(linha)
    conteudo = "\n".join(linhas)
    pyperclip.copy(conteudo)
    return conteudo

# -------------------- Funções UI -------------------- #
def atualizar_tabela():
    global checkbox_vars
    escolhido = combo.get()
    for widget in scrollable_frame.winfo_children()[len(headers):]:
        widget.destroy()

    # Limpar a lista de checkboxes
    checkbox_vars.clear()

    encontrados = [d for d in dados_planilha if d["mes"] == escolhido]
    
    # Aplicar filtro de vazias
    if filtro_var.get():
        encontrados = [d for d in encontrados if d["cobranca"] in (None, "")]
    
    # Aplicar filtro de canceladas (esconde/remove as que contêm "Cancelada")
    if filtro_canceladas_var.get():
        encontrados = [d for d in encontrados if "cancelada" in str(d.get("status", "")).lower()]

    for row_idx, d in enumerate(encontrados, start=1):
        bg_color = "#f5f5f5" if row_idx % 2 == 0 else "#ffffff"
        
        # Criar variável de checkbox e adicionar à lista (já marcado por padrão)
        var = ctk.BooleanVar(value=True)
        checkbox_vars.append({"var": var, "data": d})
        
        # Checkbox na primeira coluna - compacto
        ctk.CTkCheckBox(scrollable_frame, text="", variable=var, width=20).grid(row=row_idx, column=0, padx=2, pady=2)
        
        # Dados nas colunas seguintes
        ctk.CTkLabel(scrollable_frame, text=d["id"], bg_color=bg_color).grid(row=row_idx, column=1, padx=5, pady=2, sticky="nsew")
        ctk.CTkLabel(scrollable_frame, text=d["nome"], bg_color=bg_color).grid(row=row_idx, column=2, padx=5, pady=2, sticky="nsew")
        ctk.CTkLabel(scrollable_frame, text=d["centro"], bg_color=bg_color).grid(row=row_idx, column=3, padx=5, pady=2, sticky="nsew")
        ctk.CTkLabel(scrollable_frame, text=d.get("indice", "HRSR26"), bg_color=bg_color).grid(row=row_idx, column=4, padx=5, pady=2, sticky="nsew")
        ctk.CTkLabel(scrollable_frame, text=d.get("faturar", ""), bg_color=bg_color).grid(row=row_idx, column=5, padx=5, pady=2, sticky="nsew")
        ctk.CTkLabel(scrollable_frame, text=d.get("status", ""), bg_color=bg_color).grid(row=row_idx, column=6, padx=5, pady=2, sticky="nsew")
        ctk.CTkLabel(scrollable_frame, text=d["cobranca"], bg_color=bg_color).grid(row=row_idx, column=7, padx=5, pady=2, sticky="nsew")

filtro_var.trace_add("write", lambda *args: atualizar_tabela())
filtro_canceladas_var.trace_add("write", lambda *args: atualizar_tabela())

def executar():
    escolhido = combo.get()
    if not escolhido:
        messagebox.showwarning("Atenção", "Selecione um Mês/Ano antes de executar.")
        return

    # Filtrar apenas os itens selecionados via checkbox
    encontrados = [item["data"] for item in checkbox_vars if item["var"].get()]

    if not encontrados:
        messagebox.showinfo("Info", f"Nenhum item selecionado para {escolhido}")
        return

    preparar_clipboard(encontrados, escolhido)

    resposta = messagebox.askyesno(
        "Confirmação",
        f"Foram encontrados {len(encontrados)} registros.\n\nDeseja mesmo executar a cobrança\nreferente a {escolhido}?"
    )

    if resposta:
        abrir_sap_web(escolhido, encontrados)
    else:
        messagebox.showinfo("Cancelado", "Operação cancelada.")

def confirmar():
    atualizar_tabela()

btn_confirmar.configure(command=confirmar)
btn_cancelar_cobranca.configure(command=cancelar_cobranca_selecionados)

# Maximizar a janela após todos os widgets serem criados
app.after(0, lambda: app.state('zoomed'))

app.mainloop()
